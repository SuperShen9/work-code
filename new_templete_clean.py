# -*- coding: utf-8 -*-
# author:Super
import os,openpyxl,pprint,re
import sys
reload(sys)
sys.setdefaultencoding('utf-8')
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.styles.colors import RED
from openpyxl.styles.colors import GREEN
from openpyxl.styles.colors import BLUE
ft1 = Font(name='Arial', size=11, bold=True, color=RED)
ft2 = Font(name='Arial', size=11, bold=True, color=GREEN)
ft3 = Font(name='Arial', size=11, bold=True, color=BLUE)
os.chdir('D:\zlianxi\New_templete_clean')
wbf = openpyxl.load_workbook('Templete.xlsx')
sheet_tem = wbf.get_sheet_by_name('Sheet1')
hang_tem = sheet_tem.max_row + 1
segment={}
AM={}
job={}
dept={}
indu={}
in_pro={}
b_time={}
b_bug={}
job_level={}
for row in range(2,hang_tem):
    if sheet_tem['A' + str(row)].value != None:
        key_o = sheet_tem['A' + str(row)].value
        val_o = sheet_tem['B' + str(row)].value
        segment.setdefault(key_o, val_o)
    if sheet_tem['C' + str(row)].value != None:
        key_o = sheet_tem['C' + str(row)].value
        val_o = sheet_tem['D' + str(row)].value
        AM.setdefault(key_o, val_o)
    if sheet_tem['E' + str(row)].value != None:
        key_o = sheet_tem['E' + str(row)].value
        val_o = sheet_tem['F' + str(row)].value
        job.setdefault(key_o, val_o)
    if sheet_tem['G' + str(row)].value != None:
        key_o = sheet_tem['G' + str(row)].value
        val_o = sheet_tem['H' + str(row)].value
        dept.setdefault(key_o, val_o)
    if sheet_tem['I' + str(row)].value != None:
        key_o = sheet_tem['I' + str(row)].value
        val_o = sheet_tem['J' + str(row)].value
        indu.setdefault(key_o, val_o)
    if sheet_tem['K' + str(row)].value != None:
        key_o = sheet_tem['K' + str(row)].value
        val_o = sheet_tem['L' + str(row)].value
        in_pro.setdefault(key_o, val_o)
    if sheet_tem['N' + str(row)].value != None:
        key_o = sheet_tem['N' + str(row)].value
        val_o = sheet_tem['O' + str(row)].value
        b_time.setdefault(key_o, val_o)
    if sheet_tem['P' + str(row)].value != None:
        key_o = sheet_tem['P' + str(row)].value
        val_o = sheet_tem['Q' + str(row)].value
        b_bug.setdefault(key_o, val_o)
    if sheet_tem['R' + str(row)].value != None:
        key_o = sheet_tem['R' + str(row)].value
        val_o = sheet_tem['S' + str(row)].value
        job_level.setdefault(key_o, val_o)
os.chdir('D:\zlianxi\New_templete_clean\clean')
file = 'Clean_data.xlsx'
if os.path.exists(file):
    os.remove(file)
filepath = unicode('D:\zlianxi\New_templete_clean\clean', 'utf-8')
list1=['Segment']
list2=['AM']
list3=[u'職稱']
list4=[u'部門']
list5=[u'產業別']
list6=[u'有興趣投資的IT解決方案?(可複選)']
list7=[u'專案時程']
list8=[u'專案預算(USD)']
list9=[u'姓名']
list10=[u'完整公司名稱']
list11=[u'公司電話/分機']
list12=['Email']
list13=[u'手機']
list14=[u'地址']
list15=[u'标准職稱']
list16=[u'具體預算(USD)']
list_email=[' ',',','/','\\','@.','.@','..','，','。',':','；']
a = range(1,22)
b = list(reversed(a))
num_Regex=re.compile(r'\d+')
sub_Regex=re.compile(r'^886-|^86-')
gz02_Regex=re.compile(r'^02')
#copy
email_Regex=re.compile(r'\s|，|,|:|：|;|；|。|\||/|\\|@\.|\.@|\.\.|!|#|$|\*')
for foldername,subfolder,excels in os.walk(filepath):
    Clean_data = openpyxl.Workbook()
    sheet = Clean_data.create_sheet(index=0, title='data')
    wb=openpyxl.load_workbook(excels[0])
    sheet1 = wb.active
    hang = sheet1.max_row+1
    lie = sheet1.max_column+1
    print '数据量统计：%s'%(hang-2)
    i=0
    for k in range(1, lie):
        liebiao = get_column_letter(k)
        liebiao1=get_column_letter(k+i)
        i+=1
        for j in range(1, hang):
            sheet[liebiao1 + str(j)] = sheet1[liebiao + str(j)].value
    hang1 = sheet.max_row + 1
    lie1 = sheet.max_column + 1
    for kk in range(1 , lie1):
        lb=get_column_letter(kk)
        lb_1=get_column_letter(kk+1)
        lb_m = get_column_letter(lie1)
        for jj in range(2,hang1):
            if sheet[lb + '1'].value in list1:
                sheet[lb_1 + '1'] = '标准segment'
                sheet[lb_1 + '1'].font=ft1
                sheet[lb_1 + str(jj)] = segment.get(sheet[lb + str(jj)].value)
            if sheet[lb + '1'].value in list2:
                sheet[lb_1 + '1'] = '检验AM'
                sheet[lb_1 + '1'].font = ft1
                if sheet[lb + str(jj)].value!=None :
                    if sheet[lb + str(jj)].value in AM:
                        sheet[lb_1 + str(jj)] = AM.get(sheet[lb + str(jj)].value)
                    else:
                        sheet[lb_1 + str(jj)] ='XXXX'
            if sheet[lb + '1'].value in list3:
                sheet[lb_1 + '1'] = '标准職稱'
                sheet[lb_1 + '1'].font = ft1
                sheet[lb_1 + str(jj)] = job.get(sheet[lb + str(jj)].value)
            if sheet[lb + '1'].value in list4:
                sheet[lb_1 + '1'] = '标准部門'
                sheet[lb_1 + '1'].font = ft1
                sheet[lb_1 + str(jj)] = dept.get(sheet[lb + str(jj)].value)
            if sheet[lb + '1'].value in list5:
                sheet[lb_1 + '1'] = '标准行业'
                sheet[lb_1 + '1'].font = ft1
                sheet[lb_1 + str(jj)] = indu.get(sheet[lb + str(jj)].value)
            if sheet[lb + '1'].value in list6:
                sheet[lb_1 + '1'] = '标准产品'
                sheet[lb_1 + '1'].font = ft1
                sheet[lb_1 + str(jj)] = in_pro.get(sheet[lb + str(jj)].value)
            if sheet[lb + '1'].value in list7:
                sheet[lb_1 + '1'] = '标准时间'
                sheet[lb_1 + '1'].font = ft1
                sheet[lb_1 + str(jj)] = b_time.get(sheet[lb + str(jj)].value)
            if sheet[lb + '1'].value in list8:
                sheet[lb_1 + '1'] = '标准金额'
                sheet[lb_1 + '1'].font = ft1
                sheet[lb_1 + str(jj)] = b_bug.get(sheet[lb + str(jj)].value)
            if sheet[lb + '1'].value in list9:
                sheet[lb_1 + '1'] = '标准姓名'
                sheet[lb_1 + '1'].font = ft1
                sheet[lb_1 + str(jj)] = sheet[lb + str(jj)].value.strip()
            if sheet[lb + '1'].value in list10:
                sheet[lb_1 + '1'] = '标准公司名稱'
                sheet[lb_1 + '1'].font = ft1
                sheet[lb_1 + str(jj)] = sheet[lb + str(jj)].value.strip()
            if sheet[lb + '1'].value in list11:
                sheet[lb_1 + '1'] = '标准電話'
                sheet[lb_1 + '1'].font = ft1
                mo_sub = sub_Regex.sub('0',str(sheet[lb + str(jj)].value))
                mo_gz02 = gz02_Regex.sub('02-',mo_sub)
                mo = num_Regex.findall(mo_gz02)
                sheet[lb_1 + str(jj)] = '-'.join(mo)
    #--------------------------------------------------------------
            # if sheet[lb + '1'].value in list12:
            #     sheet[lb_1 + '1'] = '邮箱检查'
            #     sheet[lb_1 + '1'].font = ft1
            #     if sheet[lb + str(jj)].value!=None:
            #         if '@' not in sheet[lb + str(jj)].value:
            #             sheet[lb_1 + str(jj)] = '没有@'
            #         else:
            #             for l_e in list_email:
            #                 sheet[lb_1 + str(jj)] =sheet[lb + str(jj)].value.find(l_e)
            #                 if sheet[lb_1 + str(jj)].value > 0:
            #                     sheet[lb_1 + str(jj)]='出现'+l_e+'字符'
            #                     break
    # --------------------------------------------------------------
            if sheet[lb + '1'].value in list12:
                sheet[lb_1 + '1'] = '邮箱检查'
                sheet[lb_1 + '1'].font = ft1
                if sheet[lb + str(jj)].value != None:
                    if '@' not in sheet[lb + str(jj)].value:
                        sheet[lb_1 + str(jj)] = '没有@'
                    else:
                        mo1=email_Regex.findall(str(sheet[lb + str(jj)].value))
                        sheet[lb_1 + str(jj)] = '出现'+'|'.join(mo1)+'字符'

    # --------------------------------------------------------------
            if sheet[lb + '1'].value in list13:
                sheet[lb_1 + '1'] = '标准手机'
                sheet[lb_1 + '1'].font = ft1
                mo = num_Regex.findall(str(sheet[lb + str(jj)].value))
                sheet[lb_1 + str(jj)] = '/'.join(mo)
            if sheet[lb + '1'].value in list14:
                sheet[lb_1 + '1'] = '标准地址'
                sheet[lb_1 + '1'].font = ft1
                if sheet[lb + str(jj)].value!=None:
                    sheet[lb_1 + str(jj)] = sheet[lb + str(jj)].value[3:]
            if sheet[lb + '1'].value in list15:
                sheet[lb_m + '1'] = 'JOB LEVEL'
                sheet[lb_m + '1'].font = ft2
                sheet[lb_m + str(jj)] = job_level.get(sheet[lb + str(jj)].value)
# --------------------------------------------------------------------------------------------
            if sheet[lb + '1'].value in list6:
                sheet[lb_1 + '1'] = '标准产品'
                sheet[lb_1 + '1'].font = ft1
                sheet[lb_1 + str(jj)] = sheet[lb + str(jj)].value
                if sheet[lb_1 + str(jj)].value != None :
                    for cp_id in b:
                        sheet[lb_1+ str(jj)] = str(sheet[lb_1 + str(jj)].value).replace(str(cp_id),in_pro.get(cp_id))
# ------------------------------------------------------------------------------------------------
#             if sheet[lb + '1'].value in list6:
#                 sheet[lb_1 + '1'] = '标准产品'
#                 sheet[lb_1 + '1'].font = ft1
#                 sheet[lb_1 + str(jj)] = sheet[lb + str(jj)].value
#                 if sheet[lb_1 + str(jj)].value != None :
#                     sheet[lb_1 + str(jj)]=sheet[lb_1 + str(jj)].value.replace(';','|')
#                     sheet[lb_1 + str(jj)] = sheet[lb_1 + str(jj)].value.replace('；', '|')
#                     sheet[lb_1 + str(jj)] = sheet[lb_1 + str(jj)].value.replace(' ', '')
#                     sheet[lb_1 + str(jj)] = str(sheet[lb_1 + str(jj)].value).replace('雲端基礎架構與管理', 'INFRASTRUCTURE AND CLOUD MANAGEMENT')
#                     sheet[lb_1 + str(jj)] = sheet[lb_1 + str(jj)].value.replace('資料中心網路', 'DATA CENTER NETWORKING')
#                     sheet[lb_1 + str(jj)] = sheet[lb_1 + str(jj)].value.replace('資料中心伺服器', 'DATA CENTER VIRTUALIZATION - UNIFIED COMPUTING')
#                     sheet[lb_1 + str(jj)] = sheet[lb_1 + str(jj)].value.replace('網路安全', 'SECURITY - NETWORK SECURITY')
#                     sheet[lb_1 + str(jj)] = sheet[lb_1 + str(jj)].value.replace('路由器', 'ROUTERS')
#                     sheet[lb_1 + str(jj)] = sheet[lb_1 + str(jj)].value.replace('交換器', 'SWITCHES')
#                     sheet[lb_1 + str(jj)] = sheet[lb_1 + str(jj)].value.replace('無線', 'WIRELESS LAN')
#                     sheet[lb_1 + str(jj)] = sheet[lb_1 + str(jj)].value.replace('超融合基礎架構', 'CONVERGED AND HYPERCONVERGED INFRASTRUCTURE')
            if sheet[lb + '1'].value in list16:
                sheet[lb_1 + '1'] = '标准金额'
                sheet[lb_1 + '1'].font = ft3
                k_b=sheet[lb + str(jj)].value
                if k_b>=0 and k_b<1:
                    sheet[lb_1 + str(jj)] ='$0 - $999'
                elif k_b>=1 and k_b<5:
                    sheet[lb_1 + str(jj)] = '$1000 - $4999'
                elif k_b >= 5 and k_b < 15:
                    sheet[lb_1 + str(jj)] = '$5000 - $14,999'
                elif k_b>=15 and k_b<25:
                    sheet[lb_1 + str(jj)] = '$15,000 - $24,999'
                elif k_b >= 25 and k_b < 50:
                    sheet[lb_1 + str(jj)] = '$25,000 - $49,999'
                elif k_b >= 50 and k_b < 100:
                    sheet[lb_1 + str(jj)] = '$50,000 - $99,999'
                elif k_b >= 100 and k_b < 200:
                    sheet[lb_1 + str(jj)] = '$100,000-$199,999'
                elif k_b >= 200 and k_b < 500:
                    sheet[lb_1 + str(jj)] = '$200,000-$499,999'
                elif k_b >= 500 and k_b < 1000:
                    sheet[lb_1 + str(jj)] = '$500,000-$999,999'
                elif k_b >= 1000:
                    sheet[lb_1 + str(jj)] = '$1,000,000+'




sheet.freeze_panes='A2'
Clean_data.save('Clean_data.xlsx')